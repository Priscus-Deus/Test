import csv
from openpyxl.chart import BarChart, Reference
from openpyxl import Workbook, load_workbook
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows




file_path = 'Выгрузка для ЛАБ5-6.xlsx'
header_font = Font(bold=True, italic=True)
data_font = Font(bold=False)
center_aligned_text = Alignment(horizontal='center', vertical='center')
thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))



def format_date(date_object):
    
    days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
    formatted_date = date_object.strftime(f'{days[date_object.weekday()]} %m.%d.%Yг.')
    return formatted_date

def convert2xlsx():

    
    wb = Workbook()
    ws = wb.active

    
    with open('Выгрузка для ЛАБ5-6.csv', 'r', encoding='Windows-1251') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  
        for row in csvreader:
            ws.append(row)

    
    wb.save('Выгрузка для ЛАБ5-6.xlsx')



def task2():

    df = pd.read_excel('Выгрузка для ЛАБ5-6.xlsx')
    df = df.dropna(subset=[df.columns[0]])
    df['updated'] = pd.to_datetime(df['updated'], format='%d.%m.%Y %H:%M')
    full_date_range = pd.date_range(start=df['updated'].min(), end=df['updated'].max(), freq='D')
    full_df = pd.DataFrame({'updated': full_date_range})
    df = full_df.merge(df, on='updated', how='left')
    df = df.sort_values(by='updated', ascending=False)

    
    df.to_excel('Выгрузка для ЛАБ5-6.xlsx', index=False)


def task3():

    df = pd.read_excel('Выгрузка для ЛАБ5-6.xlsx')
    month_dict = {
        'янв': "01",
        'фев': "02",
        'мар': "03",
        "апр": "04",
        "май": "05",
        "июн": "06",
        "июл": "07",
        "авг": "08",
        "сен": "09",
        "окт": "10",
        "ноя": "11",
        "дек": "12"
    }

    
    def replace_month_values(value):
        
        try:
            value = str(value)
            parts = value.split('.')
            month = parts[0]
            number = parts[1] + '0'
            
            if month in month_dict:
                month_number = month_dict[month]
                return f"{month_number[1]}.{number}"
            else:
                return value
            
        except IndexError:
            pass

    # Применение функции к столбцу
    df['high'] = df['high'].apply(replace_month_values)
    df['low'] = df['low'].apply(replace_month_values)
    df['open'] = df['open'].apply(replace_month_values)
    df['close'] = df['close'].apply(replace_month_values)
    df['volume'] = df['volume'].apply(replace_month_values)
    df['average'] = df['average'].apply(replace_month_values)


    def replace_month2_values(value):
    
        try:
            value = str(value)
            parts = value.split('.')
            
            number = parts[0][1]
            month = parts[1]
            
            if month in month_dict:
                month_number = month_dict[month]
                return f"{number}.{month_number + '0'}"
            else:
                return value
        except IndexError:
            return value

    df['high'] = df['high'].apply(replace_month2_values)
    df['low'] = df['low'].apply(replace_month2_values)
    df['open'] = df['open'].apply(replace_month2_values)
    df['close'] = df['close'].apply(replace_month2_values)
    df['volume'] = df['volume'].apply(replace_month2_values)
    df['average'] = df['average'].apply(replace_month2_values)
    
    df = df.sort_values(by='updated', ascending=False)

    df = df[df['volume'] != "0.0"]
    df = df[df['volume'] != "1.0"]
    df = df[df['volume'] != "2.0"]
    mask = (df['high'] == 'None') & (df['high'].shift(1) == 'None') & (df['high'].shift(-1) == 'None')
    df = df[~mask]
    df = df[df['updated'] != '2002-06-07 00:00:00']
    
    df.replace('None', np.nan, inplace=True)

    df['high'] = df['high'].ffill()
    df['low'] = df['low'].ffill()
    df['open'] = df['open'].ffill()
    df['close'] = df['close'].ffill()
    df['volume'] = df['volume'].ffill()
    df['average'] = df['average'].ffill()
    df['average'] = df['average'].apply(lambda x: round(float(x), 3))
    df['U/R'] = df['U/R'].ffill()
    
    df.to_excel('Выгрузка для ЛАБ5-6.xlsx', index=False)

    
def task5():

    df = pd.read_excel('Выгрузка для ЛАБ5-6.xlsx')


    df_open_higher = df[df['open'] > df['close']]
    df_close_higher = df[df['close'] > df['open']]
    df['updated'] = pd.to_datetime(df['updated'])

    updated_values = df_close_higher['updated'].tolist()
    high_values = df_close_higher['high'].tolist()
    low_values = df_close_higher['low'].tolist()
    open_values = df_close_higher['open'].tolist()
    close_values = df_close_higher['close'].tolist()
    volume_values = df_close_higher['volume'].tolist()
    average_values = df_close_higher['average'].tolist()
    UR_values = df_close_higher['U/R'].tolist()
    
    wb = load_workbook("Выгрузка для ЛАБ5-6.xlsx")

    sheet_name = "task5"

    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)

    ws = wb[sheet_name]
    ws["A1"] = "updated"
    ws["B1"] = "high"
    ws["C1"] = "low"
    ws["D1"] = "open"
    ws["E1"] = "close"
    ws["F1"] = "volume"
    ws["G1"] = "average"
    ws["H1"] = "UR"
        
    for i, (updated, high, low, open, close, volume, average, UR) in enumerate(zip(updated_values, high_values, low_values, open_values, close_values, volume_values, average_values, UR_values), start=2):

        ws[f'A{i}'] = updated
        ws[f'B{i}'] = high
        ws[f'C{i}'] = low
        ws[f'D{i}'] = open
        ws[f'E{i}'] = close
        ws[f'F{i}'] = volume
        ws[f'G{i}'] = average
        ws[f'H{i}'] = UR

    updated_values = df_open_higher['updated'].tolist()
    high_values = df_open_higher['high'].tolist()
    low_values = df_open_higher['low'].tolist()
    open_values = df_open_higher['open'].tolist()
    close_values = df_open_higher['close'].tolist()
    volume_values = df_open_higher['volume'].tolist()
    average_values = df_open_higher['average'].tolist()
    UR_values = df_open_higher['U/R'].tolist()
    
    ws["J1"] = "updated"
    ws["K1"] = "high"
    ws["L1"] = "low"
    ws["M1"] = "open"
    ws["N1"] = "close"
    ws["O1"] = "volume"
    ws["P1"] = "average"
    ws["Q1"] = "UR"
        
    for i, (updated, high, low, open, close, volume, average, UR) in enumerate(zip(updated_values, high_values, low_values, open_values, close_values, volume_values, average_values, UR_values), start=2):

        ws[f'J{i}'] = updated
        ws[f'K{i}'] = high
        ws[f'L{i}'] = low
        ws[f'M{i}'] = open
        ws[f'N{i}'] = close
        ws[f'O{i}'] = volume
        ws[f'P{i}'] = average
        ws[f'Q{i}'] = UR
    

    open_higher_quarterly_volume = df_open_higher.groupby(df['updated'].dt.to_period('Q'))['volume'].sum()
    close_higher_quarterly_volume = df_close_higher.groupby(df['updated'].dt.to_period('Q'))['volume'].sum()

    df_open_higher_quarterly_volume = pd.DataFrame(open_higher_quarterly_volume)
    df_open_higher_quarterly_volume.reset_index(inplace=True)

    update_open_gisto = df_open_higher_quarterly_volume['updated'].tolist()
    volume_open_gisto = df_open_higher_quarterly_volume['volume'].tolist()

    ws["S1"] = "Q_table1"
    ws["T1"] = "volume"

    for i, (updated, volume) in enumerate(zip(update_open_gisto, volume_open_gisto), start=2):

        ws[f'S{i}'] = str(updated)
        ws[f'T{i}'] = volume
    
    chart_open = BarChart()
    
    values = Reference(ws, min_col=20, min_row=2, max_row=len(update_open_gisto))
    categories = Reference(ws, min_col=19, min_row=2, max_row=len(update_open_gisto))
    chart_open.add_data(values)
    chart_open.set_categories(categories)

    chart_open.width = 40

    ws.add_chart(chart_open, "Y2")
    
    
    
    
    
    df_close_higher_quarterly_volume = pd.DataFrame(close_higher_quarterly_volume)
    df_close_higher_quarterly_volume.reset_index(inplace=True)

    update_close_gisto = df_close_higher_quarterly_volume['updated'].tolist()
    volume_close_gisto = df_close_higher_quarterly_volume['volume'].tolist()

    ws["V1"] = "Q_table2"
    ws["W1"] = "volume"

    for i, (updated, volume) in enumerate(zip(update_close_gisto, volume_close_gisto), start=2):

        ws[f'V{i}'] = str(updated)
        ws[f'W{i}'] = volume


    


    df_close_higher_quarterly_volume = pd.DataFrame(close_higher_quarterly_volume)
    df_close_higher_quarterly_volume.reset_index(inplace=True)
    


    chart_close = BarChart()
    
    values = Reference(ws, min_col=23, min_row=2, max_row=len(update_close_gisto))
    categories = Reference(ws, min_col=22, min_row=2, max_row=len(update_close_gisto))
    chart_close.add_data(values)
    chart_close.set_categories(categories)
    chart_close.width = 40
    chart_close.height = 10
    ws.add_chart(chart_close, "Y20")

    wb.save(file_path)


def task4():
    

    wb = load_workbook("Выгрузка для ЛАБ5-6.xlsx")
    
    ws = wb["Sheet1"]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):  
        for cell in row:
            cell.value = format_date(cell.value)
    

    ws2 = wb["task5"]

    for row in ws2.iter_rows(min_row=2, max_row=2868, min_col=1, max_col=1):  
        for cell in row:
            cell.value = format_date(cell.value)


    for row in ws2.iter_rows(min_row=2, max_row=2930, min_col=10, max_col=10):  
        for cell in row:
            cell.value = format_date(cell.value)
    wb.save(file_path)


def task6():
        # Попробуем другой формат для даты, исключая день недели
    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data['Date'] = pd.to_datetime(data['updated'].str.split(' ').str[-1].str.rstrip('г.'), format='%m.%d.%Y')
    data['Year'] = data['Date'].dt.year

    # Проверим результаты преобразования снова
    #print(data[['updated', 'Date', 'Year']].head())

        # Разделим объемы торгов на децили
    data['Decile'] = pd.qcut(data['volume'], 10, labels=False)  # Нумерация децилей от 0 до 9

    # Отберем максимальные и минимальные значения для первого и десятого дециля
    min_decile_1 = data[data['Decile'] == 0]['volume'].min()
    max_decile_1 = data[data['Decile'] == 0]['volume'].max()
    min_decile_10 = data[data['Decile'] == 9]['volume'].min()
    max_decile_10 = data[data['Decile'] == 9]['volume'].max()


        # Находим годы для максимальных и минимальных значений в первом и десятом дециле
    years_min_decile_1 = data[(data['volume'] == min_decile_1) & (data['Decile'] == 0)]['Year'].unique()
    years_max_decile_1 = data[(data['volume'] == max_decile_1) & (data['Decile'] == 0)]['Year'].unique()
    years_min_decile_10 = data[(data['volume'] == min_decile_10) & (data['Decile'] == 9)]['Year'].unique()
    years_max_decile_10 = data[(data['volume'] == max_decile_10) & (data['Decile'] == 9)]['Year'].unique()


        # Создаем DataFrame для результатов
    results_df = pd.DataFrame({
        "Decile": ["1st Decile Min", "1st Decile Max", "10th Decile Min", "10th Decile Max"],
        "Volume": [min_decile_1, max_decile_1, min_decile_10, max_decile_10],
        "Years": [", ".join(years_min_decile_1.astype(str)), ", ".join(years_max_decile_1.astype(str)), 
                ", ".join(years_min_decile_10.astype(str)), ", ".join(years_max_decile_10.astype(str))]
    })

    book = load_workbook(file_path)

    # Создание нового листа в книге
    sheet_name = 'Decile Analysis'
    if sheet_name in book.sheetnames:
        del book[sheet_name]  # Удалить лист, если он уже существует
    ws = book.create_sheet(sheet_name)

    # Закрепление шапки таблицы
    ws.freeze_panes = 'A2'

    # Стилизация для заголовков и данных
    
    # Добавление заголовков с форматированием
    headers = results_df.columns
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_aligned_text
        cell.border = thin_border

    # Добавление данных с форматированием
    for r, row in enumerate(results_df.itertuples(index=False), start=2):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = data_font
            cell.alignment = center_aligned_text
            cell.border = thin_border

    # Применение жирной рамки для контура таблицы
    for row in ws['A1':f'{get_column_letter(ws.max_column)}{ws.max_row}']:
        for cell in row:
            cell.border = Border(left=Side(style='medium'), 
                                right=Side(style='medium'), 
                                top=Side(style='medium'), 
                                bottom=Side(style='medium'))

    # Сохранение изменений в файл
    book.save(file_path)



def task7():
    
    data = pd.read_excel(file_path, sheet_name='Sheet1')

    
    data['Date'] = pd.to_datetime(data['updated'].str.split(' ').str[-1].str.rstrip('г.'), format='%m.%d.%Y')
    data['Год'] = data['Date'].dt.year
    data['Месяц'] = data['Date'].dt.month

    
    data['avg_больше_close'] = data['average'] > data['close']
    data['avg_меньше_close'] = data['average'] < data['close']
    data['close_меньше_open'] = data['close'] < data['open']
    data['close_больше_open'] = data['close'] > data['open']

    
    monthly_average = data.groupby(['Год', 'Месяц'])['volume'].transform('mean')
    data['volume_up_50_percent'] = data['volume'] > 1.5 * monthly_average


    pivot_table1 = data.groupby('Год').agg({
        'avg_больше_close': 'sum',
        'avg_меньше_close': 'sum',
        'close_меньше_open': 'sum',
        'close_больше_open': 'sum'
    })

    
    pivot_table2 = data[data['volume_up_50_percent']].groupby(['Год', 'Месяц', 'U/R']).size().unstack(fill_value=0)

    pivot_table2.to_excel('f2.xlsx')
    pivot_table1.to_excel('f1.xlsx')

    df1 = pd.read_excel('f1.xlsx')
    df2 = pd.read_excel('f2.xlsx')
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        df1.to_excel(writer, sheet_name='task7_1', index=False)
        df2.to_excel(writer, sheet_name='task7_2', index=False)


def task9():
    wb = load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = 'A2'
        for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_aligned_text
                if cell.row == 1:  
                    cell.font = header_font
                else:  
                    cell.font = data_font

    wb.save(file_path)


def task8():

    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data['updated'] = pd.to_datetime(data['updated'], errors='coerce')
    data['Month'] = data['updated'].dt.to_period('M')

    monthly_averages = data.groupby('Month').mean().round(2)
    monthly_averages['U/R'] = monthly_averages['U/R'].round(0)
    monthly_averages['updated'] = monthly_averages['updated'].astype(str).str.slice(0,7)
    monthly_averages['updated'] = pd.to_datetime(monthly_averages['updated'])
    monthly_averages['updated'] = monthly_averages['updated'].dt.strftime('%B %Y')

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        monthly_averages.to_excel(writer, sheet_name='task8', index=False)



convert2xlsx()
task2()
task3()
task5()
task8()
task4()
task6()
task7()
task9()
