import openpyxl
import pandas as pd
import locale
import matplotlib.pyplot as plt
import matplotlib
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Side, Border, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

matplotlib.use('TkAgg')

month_map = {
    'янв': '1', 'фев': '2', 'мар': '3', 'апр': '4',
    'май': '5', 'июн': '6', 'июл': '7', 'авг': '8',
    'сен': '9', 'окт': '10', 'ноя': '11', 'дек': '12'
}

desired_format = "%A %d.%m.%Yг."
file_path = "result.xlsx"


def month_abr_to_value(text):
    if isinstance(text, str):
        for month, num in month_map.items():
            if month in text:
                text = text.replace(month, num)

    return text


def correct_decimal_places(value):
    if isinstance(value, float) and value < 100:
        return value

    value = str(value)
    value = value.replace('.', '')
    corrected_value = value[:1] + '.' + value[1:]
    return corrected_value


def fill_high_low_average(row):
    high = row['high']
    low = row['low']
    average = row['average']

    if pd.isna(high) and pd.isna(low):
        return row
    elif pd.isna(high) and pd.isna(average):
        return row
    elif pd.isna(low) and pd.isna(average):
        return row
    elif pd.isna(high):
        high = average - (low - average)
        row['high'] = high
    elif pd.isna(low):
        low = average - (high - average)
        row['low'] = low
    elif pd.isna(average):
        average = (high + low) / 2
        row['average'] = average

    return row


def fill_na_with_window_mean(column):
    for i in range(len(column)):
        if not pd.isna(column[i]):
            continue

        window_size = 3
        while True:
            start = max(0, i - window_size)
            end = min(len(column), i + window_size + 1)
            window_values = column[start:end]
            window_values = window_values.dropna()

            if window_values.empty:
                continue

            column[i] = round(window_values.mean())
            if column[i] != 0.5:
                break
            else:
                window_size += 1

    return column


def preprocess_data(df: pd.DataFrame):
    df.dropna(how='all', inplace=True)
    df.reset_index(drop=True, inplace=True)

    df['updated'] = pd.to_datetime(df['updated'], dayfirst=True)
    df['updated'] = df['updated'].interpolate(method='linear')

    for column in df.columns:
        if column in ['updated']:
            continue

        if column in ['U/R']:
            df[column] = fill_na_with_window_mean(df[column])
            continue

        df[column] = df[column].apply(month_abr_to_value)
        df[column] = pd.to_numeric(df[column], errors='coerce')

        if column not in ['volume']:
            df[column] = df[column].apply(correct_decimal_places)
            df[column] = pd.to_numeric(df[column], errors='coerce')

    df.loc[(df['high'] == df['low']), ['high', 'low', 'open', 'close', 'average']] = pd.NA

    df = df.apply(lambda x: fill_high_low_average(x), axis=1)

    df.loc[df['volume'] < 100, 'volume'] = pd.NA

    for column in df.columns:
        if column in ['updated', 'volume', 'U/R']:
            continue

        df[column] = df[column].interpolate(method='polynomial', order=2, limit_direction='both', limit=4)
        df[column] = df[column].round(3)

    cutoff_date = pd.to_datetime('2007-09-03')
    df['volume'] = df[df['updated'] > cutoff_date].groupby(df['updated'].dt.quarter)['volume'].transform(
        lambda x: x.fillna(round(x.mean())))

    df['updated'] = pd.to_datetime(df['updated']).dt.strftime(desired_format)

    # week_days = df['updated'].str.split().str[0]
    # print(week_days.value_counts())

    df = df[df['updated'].str.split().str[0] != 'Суббота']
    df.reset_index(drop=True, inplace=True)

    wb = openpyxl.Workbook(file_path)

    sheet_name = "data"

    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)

    ws = wb[sheet_name]

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    wb.save(file_path)

    df['updated'] = pd.to_datetime(df['updated'], format=desired_format)

    return df


def create_hist(sheet, data, cell):
    chart = BarChart()
    chart.x_axis.title = "Index"
    chart.y_axis.title = "Value"

    chart.add_data(data)

    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "2b579a"
    s.graphicalProperties.solidFill = "2b579a"

    sheet.add_chart(chart, cell)


def volume_open_close_analysis(df):
    df_higher_close: pd.DataFrame = df[df['close'] > df['open']]
    df_higher_open: pd.DataFrame = df[df['open'] > df['close']]

    df_higher_close = df_higher_close[df_higher_close['volume'].notna()]
    df_higher_open = df_higher_open[df_higher_open['volume'].notna()]

    pd.set_option('display.max_rows', None)

    print('Объемы торгов по кварталам для таблицы "Цена закрытия выше открытия":')
    df_higher_close_quart = df_higher_close.groupby(pd.Grouper(key='updated', freq='QE'))['volume'].sum()
    print(df_higher_close_quart)

    print('\nОбъемы торгов по кварталам для таблицы "Цена открытия выше закрытия":')
    print(df_higher_open.groupby(pd.Grouper(key='updated', freq='QE'))['volume'].sum())

    pd.set_option('display.max_rows', 15)

    df_higher_close['updated'] = pd.to_datetime(df_higher_close['updated'], dayfirst=True).dt.strftime(desired_format)
    df_higher_open['updated'] = pd.to_datetime(df_higher_open['updated'], dayfirst=True).dt.strftime(desired_format)

    df_higher_close = df_higher_close[['updated', 'volume']]
    df_higher_open = df_higher_open[['updated', 'volume']]

    wb = load_workbook(file_path)

    sheet_name = "5"

    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)

    ws = wb[sheet_name]

    dataframe_to_styled_rows(df_higher_close, ws)
    dataframe_to_styled_rows(df_higher_open, ws, start_col=len(df_higher_close.columns) + 2)

    data = Reference(ws, min_col=2, min_row=2, max_row=len(df_higher_close))
    create_hist(ws, data, "G2")

    data = Reference(ws, min_col=5, min_row=2, max_row=len(df_higher_open))
    create_hist(ws, data, "G20")

    wb.save(file_path)


def volume_decile_analysis(df):
    cutoff_date = pd.to_datetime('2007-09-03')
    df = df[df['updated'] > cutoff_date]

    df['decile'] = pd.qcut(df['volume'], q=10, labels=False)

    min_max_values = df.groupby('decile').agg({'volume': ['min', 'max'], 'updated': 'unique'}).reset_index()

    first_decile_max_volume = min_max_values.loc[0, ('volume', 'max')]
    tenth_decile_min_volume= min_max_values.loc[9, ('volume', 'min')]

    max_fi_de_volume_years = df[df['volume'] == first_decile_max_volume][['updated', 'volume']].drop_duplicates()
    min_te_de_volume_years = df[df['volume'] == tenth_decile_min_volume][['updated', 'volume']].drop_duplicates()

    max_fi_de_volume_years['year'] = max_fi_de_volume_years['updated'].dt.year
    min_te_de_volume_years['year'] = min_te_de_volume_years['updated'].dt.year

    max_fi_de_volume_years.drop(['updated'], axis=1, inplace=True)
    min_te_de_volume_years.drop(['updated'], axis=1, inplace=True)

    wb = load_workbook(file_path)

    sheet_name = "6"

    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)

    ws = wb[sheet_name]

    dataframe_to_styled_rows(max_fi_de_volume_years, ws)
    dataframe_to_styled_rows(min_te_de_volume_years, ws, start_col=len(max_fi_de_volume_years.columns) + 2)


    wb.save(file_path)


def dataframe_to_styled_rows(df, ws, start_row=1, start_col=1):
    header_row = list(dataframe_to_rows(df, index=False, header=True))[0]
    for j, value in enumerate(header_row, start=start_col):
        cell = ws.cell(row=start_row, column=j, value=value)
        cell.font = Font(bold=True, italic=True)
        cell.alignment = Alignment(vertical='center', horizontal='center')

    for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=1):
        for j, value in enumerate(row, start=start_col):
            cell = ws.cell(row=i + start_row, column=j, value=value)
            cell.alignment = Alignment(vertical='center', horizontal='center')

    ws.freeze_panes = ws.cell(row=2, column=start_col + len(df.columns))
    thin = Side(border_style="thin")
    for row in range(start_row, start_row + len(df) + 1):
        for col in range(start_col, start_col + len(df.columns)):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    thick = Side(border_style="thick")
    for col in range(start_col, start_col + len(df.columns)):
        ws.cell(row=start_row, column=col).border = Border(top=thick, left=thin, right=thin, bottom=thin)
        ws.cell(row=start_row + len(df), column=col).border = Border(top=thin, left=thin, right=thin, bottom=thick)

    for row in range(start_row, start_row + len(df) + 1):
        ws.cell(row=row, column=start_col).border = Border(top=thin, left=thick, right=thin, bottom=thin)
        ws.cell(row=row, column=start_col + len(df.columns) - 1).border = Border(top=thin, left=thin, right=thick, bottom=thin)

    ws.cell(row=start_row, column=start_col).border = Border(top=thick, left=thick, right=thin, bottom=thin)
    ws.cell(row=start_row, column=start_col + len(df.columns) - 1).border = Border(top=thick, left=thin, right=thick, bottom=thin)
    ws.cell(row=start_row + len(df), column=start_col).border = Border(top=thin, left=thick, right=thin, bottom=thick)
    ws.cell(row=start_row + len(df), column=start_col + len(df.columns) - 1).border = Border(top=thin, left=thin, right=thick, bottom=thick)


def summaries(df):
    summary_df = pd.DataFrame(columns=['Year', 'Avg > Close', 'Avg < Close', 'Close < Open', 'Close > Open'])

    for year in df['updated'].dt.year.unique():
        year_df = df[df['updated'].dt.year == year]
        avg_gt_close = (year_df['average'] > year_df['close']).sum()
        avg_lt_close = (year_df['average'] < year_df['close']).sum()
        close_lt_open = (year_df['close'] < year_df['open']).sum()
        close_gt_open = (year_df['close'] > year_df['open']).sum()

        days_with_high_volume = 0
        for month in year_df['updated'].dt.month.unique():
            month_df = year_df[year_df['updated'].dt.month == month]
            monthly_avg_volume = month_df['volume'].mean()
            days_with_high_volume += (month_df['volume'] > monthly_avg_volume).sum()

        summary_df = pd.concat([summary_df, pd.DataFrame(
            {'Year': [year], 'Avg > Close': [avg_gt_close], 'Avg < Close': [avg_lt_close],
             'Close < Open': [close_lt_open], 'Close > Open': [close_gt_open]})])

    volume_df = pd.DataFrame(columns=['Month', 'Days with Volume > Monthly Avg'])

    for month in df['updated'].dt.month.unique():
        month_df = df[df['updated'].dt.month == month]
        monthly_avg_volume = month_df['volume'].mean()
        days_with_high_volume = (month_df['volume'] > monthly_avg_volume).sum()

        volume_df = pd.concat(
            [volume_df, pd.DataFrame({'Month': [month], 'Days with Volume > Monthly Avg': [days_with_high_volume]})])

    volume_df.sort_values(by=['Month'], inplace=True)

    wb = load_workbook(file_path)

    sheet_name = "7"

    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)

    ws = wb[sheet_name]

    dataframe_to_styled_rows(summary_df, ws)
    dataframe_to_styled_rows(volume_df, ws, start_col=len(summary_df.columns) + 2)

    wb.save(file_path)


def month_aggregation(df):
    monthly_data = df.groupby(pd.Grouper(key='updated', freq='ME')).agg({
        'high': 'mean',
        'low': 'mean',
        'open': 'mean',
        'close': 'mean',
        'volume': 'mean',
        'average': 'mean',
        'U/R': 'mean'
    }).reset_index()

    for column in monthly_data.columns:
        if column in ['updated']:
            monthly_data[column] = pd.to_datetime(monthly_data['updated']).dt.strftime("%m.%Y")
            continue

        if column in ['volume']:
            monthly_data[column] = monthly_data[column].round()
            continue

        monthly_data[column] = monthly_data[column].round(3)

    wb = load_workbook(file_path)

    sheet_name = "8"

    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)

    ws = wb[sheet_name]

    dataframe_to_styled_rows(monthly_data, ws)

    wb.save(file_path)


def main():
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', 1000)
    pd.options.mode.chained_assignment = None
    locale.setlocale(locale.LC_TIME, 'ru_RU.UTF-8')

    df = pd.read_csv("Выгрузка для ЛАБ5-6.csv", sep=';', encoding='windows-1251')
    df = preprocess_data(df)

    print(df)
    print(df.dtypes)

    volume_open_close_analysis(df)

    volume_decile_analysis(df)

    summaries(df)

    month_aggregation(df)


if __name__ == '__main__':
    main()