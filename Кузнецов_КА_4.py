
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
import pandas as pd
from openpyxl.drawing.colors import ColorChoice
import statistics
from openpyxl.styles import PatternFill
   

file_path = 'Статистика.xlsx'

months = {
        'Январь': 1,
        'Февраль': 2,
        'Март': 3,
        'Апрель': 4,
        'Май': 5,
        'Июнь': 6,
        'Июль': 7,
        'Август': 8,
        'Сентябрь': 9,
        'Октябрь': 10,
        'Ноябрь': 11,
        'Декабрь': 12
    }

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

print("Программа, помогающая совершать аналитические действия над файлом Excel")
print("Выберете действия")
print("1 - Анализ в разрезе отдельного месяца за несколько лет")
print("2 - Анализ в разрезе изменения показателя за все месяцы указанного промежутка лет")
print("0 - Завершение работы программы")



while True:
    menu = int(input("Сделайте выбор -> "))

    if menu == 0:
        break

    elif menu == 1:
        month = input("Введите название месяца в именительном падеже -> ").capitalize()
        start, end = map(int,input("Введите интервал в годах в формате ХХХХ-ХХХХ -> ").split("-"))

        rd_table = pd.read_excel(file_path)

        col1 = rd_table['Месяц']
        col2 = rd_table['Цена, USD']

        years = []
        prices = []

        for i in  rd_table.index:
            if col1[i].split(' ')[0] == month and int(col1[i].split(' ')[1]) >= start and int(col1[i].split(' ')[1]) <= end:
                years.append(int(col1[i].split(' ')[1]))
                prices.append(float(col2[i]))
                

        wb = load_workbook(file_path)

        sheet_name = month + "_" + str(start) + "_" + str(end)

        if sheet_name not in wb.sheetnames:
            wb.create_sheet(title=sheet_name)

        ws = wb[sheet_name]

        ws['A1'] = month

        ws['A2'] = "Год"
        ws['B2'] = "Цена, USD"

        for i, (year, price) in enumerate(zip(years, prices), start=3):

            ws[f'A{i}'] = year
            ws[f'B{i}'] = price



        chart = ScatterChart()
        

        x_values = Reference(ws, min_col=1, min_row=3, max_row=ws.max_row)
        y_values = Reference(ws, min_col=2, min_row=3, max_row=ws.max_row)

        series = Series(y_values, x_values, title='Цена, USD')
        chart.series.append(series)
        series.graphicalProperties.line.solidFill = ColorChoice(prstClr="black")

        chart.title = "Цены по годам"
        chart.x_axis.title = "Год"
        chart.y_axis.title = "Цена, USD"

        chart.style = 13

        ws.add_chart(chart, "E5")

        wb.save(file_path)

       

    elif menu == 2:
        
        start_month, start_year = map(int,input("Введите начало желаемого промежутка в формате MM_YYYY -> ").split("_"))
        end_month, end_year = map(int,input("Введите конец желаемого промежутка в формате MM_YYYY -> ").split("_"))

        rd_table = pd.read_excel(file_path)

        col1 = rd_table['Месяц']
        col2 = rd_table['Цена, USD']

        df_months = []
        df_price = []

        for i in rd_table.index:
            
            if int(col1[i].split(" ")[1]) > start_year and int(col1[i].split(" ")[1]) < end_year:
                df_months.append(col1[i])
                df_price.append(col2[i])

            elif int(col1[i].split(" ")[1]) == start_year:
                
                if months[col1[i].split(" ")[0]] >= start_month:
                    df_months.append(col1[i])
                    df_price.append(col2[i])

            elif int(col1[i].split(" ")[1]) == end_year:

                if months[col1[i].split(" ")[0]] <= end_month:
                    df_months.append(col1[i])
                    df_price.append(col2[i])

        wb = load_workbook(file_path)

        sheet_name = "анализ"+ "_" + str(start_month) + "_" + str(start_year) + "_" + str(end_month) + "_" + str(end_year)

        if sheet_name not in wb.sheetnames:
            wb.create_sheet(title=sheet_name)

        ws = wb[sheet_name]

        avg = sum(col2) / len(col2)

        standard_deviation = statistics.stdev(col2)

        ws['A2'] = "Год"
        ws['B2'] = "Цена, USD"
        ws['D2'] = 'Среднее значение стоимости газа за указанный период'
        ws['D3'] = avg
        ws['E2'] = 'среднее квадратическое отклонение по ряду'
        ws['E3'] = standard_deviation


        for i, (year, price) in enumerate(zip(df_months, df_price), start=3):

            ws[f'A{i}'] = year
            ws[f'B{i}'] = price


        for i in range(3, ws.max_row + 1):
            if (abs(avg - float(ws[f'B{i}'].value)) / standard_deviation) < 1:
                ws[f'B{i}'].fill = green_fill
            
            elif (abs(avg - float(ws[f'B{i}'].value)) / standard_deviation) > 1 and (abs(avg - float(ws[f'B{i}'].value)) / standard_deviation) < 3:
                ws[f'B{i}'].fill = yellow_fill
            
            elif (abs(avg - float(ws[f'B{i}'].value)) / standard_deviation) > 3:
                ws[f'B{i}'].fill = red_fill

        wb.save(file_path)

    else:
        print("Введены недопустимые значения")



